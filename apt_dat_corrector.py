import os
from openpyxl import load_workbook
import unicodedata

# ---------------- CONFIG ----------------
CUSTOM_SCENERY = r"C:\Users\Daniel\Desktop\X-Plane 12\Custom Scenery"
XLSX_PATH = r"C:\Users\Daniel\Downloads\airports.xlsx"
DRY_RUN = False
VERBOSE = True
# ----------------------------------------


def strip_accents(text):
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def load_airport_map_xlsx(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    airport_map = {}

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 3:
            continue

        icao = row[1]
        name = row[2]

        if not icao or not name:
            continue

        icao = str(icao).strip().upper()
        name = strip_accents(str(name).strip())

        airport_map[icao] = name

    return airport_map


def process_apt_dat(apt_path, airport_map, dry_run=True):
    try:
        with open(apt_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except Exception as e:
        print(f"[ERROR] Cannot read {apt_path}: {e}")
        return

    if len(lines) < 4:
        return

    line4 = lines[3].strip()
    parts = line4.split()

    # MUST have at least:
    # 1 5354 1 0 ICAO NAME...
    if len(parts) < 6:
        return

    icao = parts[4].upper()
    old_name = " ".join(parts[5:]).strip()

    if icao not in airport_map:
        print(f"SKIP (no ICAO match): {icao}")
        return

    new_name = airport_map[icao].strip()

    if new_name == old_name:
        print(f"SKIP (no change): {icao}")
        return

    print(f"{icao} {old_name} -> {icao} {new_name}")

    if not dry_run:
        lines[3] = f"{parts[0]} {parts[1]} {parts[2]} {parts[3]} {icao} {new_name}\n"

        try:
            with open(apt_path, "w", encoding="utf-8", errors="replace") as f:
                f.writelines(lines)
        except Exception as e:
            print(f"[ERROR WRITING] {apt_path}: {e}")


def main():
    print("Loading Excel airport data...")
    airport_map = load_airport_map_xlsx(XLSX_PATH)
    print(f"Loaded {len(airport_map)} airports\n")

    print("Scanning Custom Scenery...\n")

    for root, _, files in os.walk(CUSTOM_SCENERY):
        for file in files:
            if file.lower() == "apt.dat":
                process_apt_dat(os.path.join(root, file), airport_map, DRY_RUN)

    print("\nDone.")


if __name__ == "__main__":
    main()