"""Core functionality for apt.dat corrector"""

import os
import unicodedata
from typing import Callable, Optional, Dict, Tuple
from openpyxl import load_workbook


def strip_accents(text):
    """Remove accent marks from text"""
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def load_airport_map_xlsx(xlsx_path: str, log_fn: Optional[Callable] = None) -> Dict:
    """Load airport map from Excel file"""
    def log(msg):
        if log_fn:
            log_fn(msg)
    
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        airport_map = {}

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 3:
                continue

            icao = row[1]
            name = row[2]

            if not icao or not name:
                continue

            icao = str(icao).strip().upper()
            name = strip_accents(str(name).strip())

            airport_map[icao] = name

        log(f"Loaded {len(airport_map)} airports from Excel\n")
        return airport_map
    except Exception as e:
        log(f"Error loading Excel: {e}\n")
        return {}


def process_apt_dat(apt_path: str, airport_map: Dict, dry_run: bool = True,
                   log_fn: Optional[Callable] = None) -> Optional[Tuple[str, str, str]]:
    """Process a single apt.dat file. Returns (icao, old_name, new_name) if changed, else None."""
    def log(msg):
        if log_fn:
            log_fn(msg)

    try:
        with open(apt_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except Exception as e:
        log(f"  Cannot read {apt_path}: {e}")
        return None

    if len(lines) < 4:
        log(f"  File too small: {apt_path}")
        return None

    line4 = lines[3].strip()
    parts = line4.split()

    if len(parts) < 6:
        log(f"  Invalid format: {apt_path}")
        return None

    icao = parts[4].upper()
    old_name = " ".join(parts[5:]).strip()

    if icao not in airport_map:
        log(f"  SKIP (no ICAO): {icao}")
        return None

    new_name = airport_map[icao].strip()

    if new_name == old_name:
        log(f"  SKIP (no change): {icao}")
        return None

    log(f"  UPDATE: {icao}")
    log(f"    OLD: {old_name}")
    log(f"    NEW: {new_name}\n")

    if not dry_run:
        lines[3] = f"{parts[0]} {parts[1]} {parts[2]} {parts[3]} {icao} {new_name}\n"

        try:
            with open(apt_path, "w", encoding="utf-8", errors="replace") as f:
                f.writelines(lines)
        except Exception as e:
            log(f"    ERROR WRITING: {e}")
            return None

    return (icao, old_name, new_name)


def scan_and_process(scenery_path: str, airport_map: Dict, dry_run: bool = True,
                    log_fn: Optional[Callable] = None) -> Dict:
    """Scan directory and process all apt.dat files"""
    def log(msg):
        if log_fn:
            log_fn(msg)

    stats = {"total": 0, "changed": 0, "skipped": 0, "errors": 0, "changes": []}

    if not os.path.isdir(scenery_path):
        log(f"Directory not found: {scenery_path}\n")
        return stats

    log(f"Scanning: {scenery_path}\n")

    for root, _, files in os.walk(scenery_path):
        for file in files:
            if file.lower() == "apt.dat":
                stats["total"] += 1
                full_path = os.path.join(root, file)
                rel_path = os.path.relpath(full_path, scenery_path)
                log(f"[{stats['total']}] {rel_path}\n")

                change = process_apt_dat(full_path, airport_map, dry_run, log_fn)
                if change:
                    stats["changed"] += 1
                    stats["changes"].append(change)
                else:
                    stats["skipped"] += 1

    log("\n" + "="*60)
    log(f"RESULTS:")
    log(f"  Total files: {stats['total']}")
    log(f"  Changed: {stats['changed']}")
    log(f"  Skipped: {stats['skipped']}")
    log(f"  Dry run: {dry_run}")
    log("="*60 + "\n")

    return stats
